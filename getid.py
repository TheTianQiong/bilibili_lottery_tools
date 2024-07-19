# 网络工具 爬取数据
import datetime
import json
import os
import re

import openpyxl
import requests
from colorama import Fore


class LikeAndForwardOperation:  # v1.0
    def __init__(self, id: str, cookie: str):
        """

        :param id: 是每个动态的id，通常可以在动态页面的url中找到
        :param cookie: 是用户登录后请求内必有内容，若缺省则会导致获取不到内容。一般从开发者工具里请求的cookie中直接复制
        :return:

        """
        global end_message
        offset = ""  # 初始为空表示第一页
        page = 1
        lst = []
        url = "https://api.bilibili.com/x/polymer/web-dynamic/v1/detail/reaction"
        headers = {
            "Cookie": cookie,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0"
        }
        params = {"id": id, "offset": offset}
        begin_message, plus_value = 1, 0
        count = 0
        self.wb_new()
        while True:
            r = requests.get(url, params=params, headers=headers).json()
            total = json.loads(json.dumps(r))["data"]["total"]
            page_total = total // 20 if total % 20 == 0 else total // 20 + 1
            print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 正在获取第{page}/{page_total}页的内容中……",
                  end="")
            dict_r = json.loads(json.dumps(r))
            lst = lst + dict_r["data"]["items"]
            plus_value = plus_value + len(dict_r["data"]["items"])
            count = count + len(dict_r["data"]["items"])
            print(f"已合并")
            if not dict_r["data"]['has_more']:
                break
            if page % 20 == 0:  # 防止数据量较大，每到20页存档一次，并清空
                self.wb_append(lst, begin_message + 1)
                lst = []
                begin_message = begin_message + plus_value
                plus_value = 0
            offset = dict_r["data"]["offset"]
            params = {"id": id, "offset": offset}
            page += 1
        print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 共计获取到{count}条数据，给定数据量为{total}")
        if count != total:
            print(
                Fore.RED + "注意：这里获取到的数据量和给定的数据量不一致，可能爬取的是非本UP主的且数据量较大的动态，\n若想要完整的获取，请登录本人的账号再试；也可能是存在新的数据而导致的" + Fore.RESET)
        self.wb_append(lst, begin_message + 1)
        print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 数据处理完成")

    def wb_new(self):
        print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 新建LikeAndForward.xlsx文件……", end="")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        ws['A1'], ws['B1'], ws['C1'], ws['D1'] = "Name", "Mid", "Action", "Face"
        try:
            wb.save('LikeAndForward.xlsx')
        except:
            print(Fore.RED + "此处无法新建文档，可能是文件被占用，请关闭可能的程序" + Fore.RESET)
        print("完成")

    def wb_append(self, lst, start):
        print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 打开LikeAndForward.xlsx文件")
        wb = openpyxl.load_workbook('LikeAndForward.xlsx')
        ws = wb.active
        w = start
        print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 正在存入目前内容……", end="")
        for i in lst:
            ws["A" + str(w)], ws["B" + str(w)], ws["C" + str(w)], ws["D" + str(w)] = i["name"], i["mid"], i["action"], \
                i["face"]
            w += 1
        wb.save('LikeAndForward.xlsx')
        print("完成")


class CommentsOperation:  # v1.1
    def __init__(self, id, cookie):
        """

                :param id: 是每个动态的oid，通常可以在动态页面的url中找到，视频可以直接输入BVid（一定要有BV）
                :param cookie: 是用户登录后请求内必有内容，若缺省则会导致获取不到内容。一般从开发者工具里请求的cookie中直接复制
                :return:
                """
        print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 获取oid和type")
        if "BV" in id:
            (oid, type) = self.bv_oid_get(id, cookie)
        else:
            (oid, type) = self.opus_oid_get(id, cookie)
        global page
        self.lst, self.Ccount, startComment = [], 0, 1
        url = "https://api.bilibili.com/x/v2/reply"
        params = {"type": type, "oid": oid, "nohot": 1}
        headers = {
            "Cookie": cookie,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0"
        }
        print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 正在获取第1页的内容中……", end="")
        r = requests.get(url, params=params, headers=headers).json()
        dict_r = json.loads(json.dumps(r))
        print("已完成")
        if dict_r["code"] != 0:
            print(Fore.RED + dict_r["message"] + Fore.RESET)
            if dict_r["code"] == -400:
                print(Fore.RED + "请求错误" + Fore.RESET)
            if dict_r["code"] == -404:
                print(Fore.RED + "无此项" + Fore.RESET)
            if dict_r["code"] == 12002:
                print(Fore.RED + "评论区已关闭" + Fore.RESET)
            if dict_r["code"] == 12009:
                print(Fore.RED + "评论主体的type不合法" + Fore.RESET)
        else:
            self.wb_new()
            count = dict_r["data"]["page"]["count"]
            page = 1
            page_total = count // 20 if count % 20 == 0 else count // 20 + 1
            print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 正在处理第{page}/{page_total}页的内容中……",
                  end="")
            replies_list = dict_r["data"]["replies"]
            print("已完成")
            for rep in replies_list:  # 循环
                if rep["count"] != 0:
                    self.cycle_get_comment_id(rep["replies"])
                self.Ccount = self.Ccount + 1
                self.lst.append(self.common_dict_update(rep))
            try:
                self.wb_append(self.lst, startComment + 1)
                self.lst = []
                startComment = startComment + 20
                # 判断是否为一页，再决定是否进入循环
                if page_total != 1:
                    for page in range(2, page_total + 1):
                        params["pn"] = page
                        # 请求
                        print(
                            f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 正在获取第{page}/{page_total}页的内容中……",
                            end="")
                        r = requests.get(url, params=params, headers=headers).json()
                        dict_r = json.loads(json.dumps(r))
                        print("已完成")
                        print(
                            f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 正在处理第{page}/{page_total}页的内容并加入文件中……",
                            end="")
                        replies_list = dict_r["data"]["replies"]
                        print("已完成")
                        for rep in replies_list:
                            if rep["count"] != 0:
                                self.cycle_get_comment_id(rep["replies"])
                            self.Ccount = self.Ccount + 1
                            self.lst.append(self.common_dict_update(rep))
                        self.wb_append(self.lst, startComment + 1)
                        self.lst = []
                        startComment = startComment + 20
                # 结束循环提示
                print(
                    f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 完成！共处理{page_total}页评论，共{count}条评论")
                if count != self.Ccount:
                    print(
                        Fore.RED + "注意：这里获取到的数据量和给定的数据量不一致，可能爬取的是非本UP主的且数据量较大的动态，\n若想要完整的获取，请登录本人的账号再试；也可能是存在新的数据而导致的&或者评论被删除" + Fore.RESET)
            except:
                print(Fore.RED + "此处无法写入文档，可能是文件被占用，请关闭可能的程序" + Fore.RESET)

    def wb_new(self):
        print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 新建Comments.xlsx文件……", end="")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'], ws['F1'], ws['G1'], ws['H1'], ws['I1'], ws[
            'J1'] = "Name", "Mid", "Sex", "Face", "Level", "Fansgrade", "Message", "Oid", "Rpid", "Root"
        try:
            wb.save('Comments.xlsx')
        except:
            print(Fore.RED + "此处无法新建文档，可能是文件被占用，请关闭可能的程序" + Fore.RESET)
        print("完成")

    def wb_append(self, lst, start):
        print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 打开Comments.xlsx文件")
        wb = openpyxl.load_workbook('Comments.xlsx')
        ws = wb.active
        w = start
        print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} 正在存入目前内容……", end="")
        for i in lst:
            ws["A" + str(w)], ws["B" + str(w)], ws["C" + str(w)], ws["D" + str(w)], ws["E" + str(w)], ws["F" + str(w)], \
                ws["G" + str(w)], ws["H" + str(w)], ws["I" + str(w)], ws["J" + str(w)] = i["name"], i["mid"], i["sex"], \
            i[
                "face"], i["level"], i["fansgrade"], i["message"], i["oid"], i["rpid"], i["root"]
            w += 1
        wb.save('Comments.xlsx')
        print("完成")

    def common_dict_update(self, rep):
        dict_rep = {}
        dict_rep["oid"] = rep["oid"]
        dict_rep["fansgrade"] = rep["fansgrade"]
        dict_rep["mid"] = rep["member"]["mid"]
        dict_rep["name"] = rep["member"]["uname"]
        dict_rep["sex"] = rep["member"]["sex"]
        dict_rep["face"] = rep["member"]["avatar"]
        dict_rep["level"] = rep["member"]["level_info"]["current_level"]
        dict_rep["message"] = rep["content"]["message"]
        dict_rep["root"] = rep["root"]
        dict_rep["rpid"] = rep["rpid"]
        return dict_rep

    def opus_oid_get(self, id, cookie):
        url = "https://www.bilibili.com/opus/" + str(id)
        headers = {
            "Cookie": cookie,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0"
        }
        reponse = requests.get(url, headers=headers)
        # print(reponse.text)
        comment_id_str = re.search(r'"comment_id_str":"(.*?)"', reponse.text, re.M).group(1)
        comment_type = re.search(r'"comment_type":(.*?),', reponse.text, re.M).group(1)
        return (comment_id_str, comment_type)

    def bv_oid_get(self, id, cookie):
        url = "https://www.bilibili.com/video/" + str(id)
        headers = {
            "Cookie": cookie,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0"
        }
        reponse = requests.get(url, headers=headers)
        # print(reponse.text)
        aid = re.search(r'"aid":(.*?),"bvid":"(.*?)"', reponse.text, re.M).group(1)
        type = 1
        return (aid, type)

    def cycle_dict_update(self, rep):
        dict_rep = {}
        dict_rep["oid"] = rep["oid"]
        dict_rep["fansgrade"] = rep["fansgrade"]
        dict_rep["mid"] = rep["mid"]
        dict_rep["name"] = rep["name"]
        dict_rep["sex"] = rep["sex"]
        dict_rep["face"] = rep["face"]
        dict_rep["level"] = rep["level"]
        dict_rep["message"] = rep["message"]
        dict_rep["root"] = rep["root"]
        dict_rep["rpid"] = rep["rpid"]
        return dict_rep

    def cycle_get_comment_id(self, Clst):
        for rep in Clst:
            # print(rep)
            if rep["count"] != 0:
                self.cycle_get_comment_id(rep["replies"])
            self.Ccount = self.Ccount + 1
            if "name" in rep:
                self.lst.append(self.cycle_dict_update(rep))
            else:
                self.lst.append(self.common_dict_update(rep))


if __name__ == '__main__':
    print(
        "本程序已在github上开源，地址为" + Fore.BLUE + "https://github.com/TheTianQiong/bilibili_lottery_tools" + Fore.RESET + "\n受GPL-3开源协议保护，不得商用！" + Fore.BLUE + "联系：1214853517@qq.com" + Fore.RESET + "\n版本号：v1.1")
    choice = input("请选择获取的内容：" + Fore.GREEN + "（1:点赞和转发；2:评论）" + Fore.RESET)
    if choice == "1":
        print(Fore.YELLOW + "视频请使用发送动态的id" + Fore.RESET)
        id = str(input("请输入要获取的动态id号："))
        cookie = str(input("请输入cookie值："))
        LikeAndForwardOperation(id, cookie)
    elif choice == "2":
        print(Fore.YELLOW + "视频一定要输入BVid(包含'BV'字符)，否则会报错" + Fore.RESET)
        id = str(input("请输入要获取的id号："))
        cookie = str(input("请输入cookie值："))
        CommentsOperation(id, cookie)
    os.system('pause')
